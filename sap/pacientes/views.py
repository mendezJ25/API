from django.forms import modelform_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.template import loader
from openpyxl.workbook import Workbook

from pacientes.forms import PacienteFormulario
from pacientes.models import Paciente

# Create your views here.


def agregar_paciente(request):
    pagina = loader.get_template('agregar_pacientes.html')
    if request.method == 'GET':
         formulario = PacienteFormulario
    elif request.method == 'POST':
        formulario = PacienteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_paciente(request, idPaciente):
    pagina = loader.get_template('ver_paciente.html')
    # persona = Persona.objects.get(pk=idPersona)
    paciente = get_object_or_404(Paciente, pk=idPaciente)
    mensaje = {'paciente': paciente}
    return HttpResponse(pagina.render(mensaje, request))


def editar_paciente(request, idPaciente):
    pagina = loader.get_template('editar_paciente.html')
    paciente = get_object_or_404(Paciente, pk=idPaciente)
    if request.method == "GET":
        formulario = PacienteFormulario(instance=paciente)

    elif request.method == "POST":
        formulario = PacienteFormulario(request.POST, instance=paciente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    mensaje = {'formulario': formulario}

    return HttpResponse(pagina.render(mensaje, request))

def eliminar_paciente(request, idPaciente):
    paciente = get_object_or_404(Paciente, pk=idPaciente)
    if paciente:
        paciente.delete()
        return redirect('inicio')



def generar_reporte(request):
    # pacientes = Paciente.objects.all()
    pacientes = Paciente.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE PACIENTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'EDAD'
    ws['E3'] = 'SEXO'
    ws['F3'] = 'EMAIL'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for paciente in pacientes:
        ws.cell(row=cont, column=2).value = paciente.nombre
        ws.cell(row=cont, column=3).value = paciente.apellido
        ws.cell(row=cont, column=4).value = paciente.edad
        ws.cell(row=cont, column=5).value = paciente.sexo
        ws.cell(row=cont, column=6).value = paciente.email
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReportePacientesExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
